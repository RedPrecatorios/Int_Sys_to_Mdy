"""
Cria (ou sobrescreve) uma aba no Excel FINAL com Telefone, Nome e Número de processo.

Dois layouts suportados (detecção automática pelo cabeçalho):

1) **lemitti_uvz** — export Lemitti “puro”:
   - DDD coluna U, telefone coluna V → dígitos(U)+dígitos(V), sem 55
   - Coluna Z POSSUI-WHATSAPP = 1

2) **prc_final** — arquivo tipo `PRC TJSP FINAL.xlsx`:
   - Colunas TELEFONE_1 … TELEFONE_n
   - Se existirem POSSUI_WHATSAPP_1 … (ou POSSUI-WHATSAPP-1 etc.), usa só telefone com flag = 1
   - Se não houver colunas de flag, usa o primeiro TELEFONE_k preenchido (primeiro contato),
     a menos que `exige_flag_whatsapp=True` (aí a linha é ignorada)

Nome e processo: cabeçalhos (ex.: Requerente, Numero_de_Processo) ou colunas informadas por letra.

3) **Dois ficheiros** (`caminho_enriquecimento_lemitti` preenchido):
   - Ficheiro Lemitti (enriquecimento): colunas **U (DD) + V (telefone)** → número sem 55; **Z = 1** → WhatsApp.
   - Ficheiro FINAL: lê **Requerente** e **Numero_de_Processo** (aba tipo PRC / Sheet1).
   - Cruza pelo **número de processo** normalizado; um telefone por processo (primeira linha Lemitti válida).
"""

from __future__ import annotations

import re
import unicodedata
from dataclasses import dataclass
from typing import Dict, Iterator, List, Literal, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet

Layout = Literal["lemitti_uvz", "prc_final"]

COL_DDD = column_index_from_string("U")
COL_TEL = column_index_from_string("V")
COL_WA = column_index_from_string("Z")

NOME_ABA_DESTINO_PADRAO = "WhatsApp_FINAL"


def _only_digits(s: object) -> str:
    if s is None:
        return ""
    return re.sub(r"\D", "", str(s).strip())


def _whatsapp_flag(val: object) -> bool:
    if val is None or val == "":
        return False
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return int(val) == 1
    t = str(val).strip().upper()
    return t in ("1", "SIM", "TRUE", "S", "YES")


def _norm_header(s: str) -> str:
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", str(s))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"[^a-z0-9]", "", s.lower())


def _header_map(ws: Worksheet, header_row: int) -> Dict[str, int]:
    """Título normalizado → índice de coluna (1-based)."""
    m: Dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col).value
        if cell is None:
            continue
        m[_norm_header(str(cell))] = col
    return m


def _detect_layout(ws: Worksheet, header_row: int) -> Optional[Layout]:
    hm = _header_map(ws, header_row)
    # PRC FINAL: cabeçalho TELEFONE_1 → chave normalizada "telefone1"
    if "telefone1" in hm:
        return "prc_final"
    z_key = _norm_header(str(ws.cell(row=header_row, column=COL_WA).value or ""))
    if "possui" in z_key and "what" in z_key:
        return "lemitti_uvz"
    return None


def _find_header_columns(
    ws: Worksheet,
    header_row: int,
) -> Tuple[Optional[int], Optional[int]]:
    nome_idx: Optional[int] = None
    proc_idx: Optional[int] = None

    nome_tokens = (
        "nome",
        "requerente",
        "nomecompleto",
        "nome_completo",
        "credor",
    )
    proc_tokens = (
        "processo",
        "numerodoprocesso",
        "numero_processo",
        "nrprocesso",
        "processoprincipal",
        "elemento",
    )

    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col).value
        if cell is None:
            continue
        h = _norm_header(str(cell))
        if nome_idx is None and any(t in h for t in nome_tokens):
            nome_idx = col
        if proc_idx is None and any(t in h for t in proc_tokens):
            proc_idx = col

    return nome_idx, proc_idx


def _parse_telefone_possui_cols(ws: Worksheet, header_row: int) -> Tuple[Dict[int, int], Dict[int, int]]:
    """
    Retorna (telefone_por_indice, possui_whatsapp_por_indice).
    Aceita cabeçalhos: TELEFONE_1, TELEFONE-1, POSSUI_WHATSAPP_1, POSSUI-WHATSAPP-1, etc.
    """
    telefones: Dict[int, int] = {}
    possui: Dict[int, int] = {}

    for col in range(1, ws.max_column + 1):
        raw = ws.cell(row=header_row, column=col).value
        if raw is None:
            continue
        title = str(raw).strip().upper().replace(" ", "_")
        m = re.match(r"^TELEFONE[_-]?(\d+)$", title)
        if m:
            telefones[int(m.group(1))] = col
            continue
        m = re.match(r"^POSSUI[_-]?WHATSAPP[_-]?(\d+)$", title)
        if m:
            possui[int(m.group(1))] = col
            continue

    return telefones, possui


def _pick_telefone_prc_final(
    ws: Worksheet,
    row: int,
    telefones: Dict[int, int],
    possui: Dict[int, int],
    exige_flag: bool,
) -> str:
    """Primeiro telefone válido conforme regras prc_final."""
    for k in sorted(telefones.keys()):
        tel_col = telefones[k]
        digits = _only_digits(ws.cell(row=row, column=tel_col).value)
        if not digits or len(digits) < 8:
            continue
        if k in possui:
            if not _whatsapp_flag(ws.cell(row=row, column=possui[k]).value):
                continue
        elif exige_flag:
            continue
        return digits
    return ""


def iter_linhas_lemitti_uvz(
    ws: Worksheet,
    data_start: int,
    col_nome: int,
    col_processo: int,
) -> Iterator[Tuple[str, str, str]]:
    for row in range(data_start, ws.max_row + 1):
        ddd = _only_digits(ws.cell(row=row, column=COL_DDD).value)
        tel = _only_digits(ws.cell(row=row, column=COL_TEL).value)
        flag = ws.cell(row=row, column=COL_WA).value

        if not _whatsapp_flag(flag):
            continue
        if not ddd and not tel:
            continue

        telefone = f"{ddd}{tel}"
        if not telefone:
            continue

        nome = ws.cell(row=row, column=col_nome).value
        proc = ws.cell(row=row, column=col_processo).value
        nome_s = str(nome).strip() if nome is not None else ""
        proc_s = str(proc).strip() if proc is not None else ""

        yield telefone, nome_s, proc_s


def iter_linhas_prc_final(
    ws: Worksheet,
    header_row: int,
    data_start: int,
    col_nome: int,
    col_processo: int,
    exige_flag_whatsapp: bool,
) -> Iterator[Tuple[str, str, str]]:
    telefones, possui = _parse_telefone_possui_cols(ws, header_row)
    if not telefones:
        return
    for row in range(data_start, ws.max_row + 1):
        telefone = _pick_telefone_prc_final(ws, row, telefones, possui, exige_flag_whatsapp)
        if not telefone:
            continue
        nome = ws.cell(row=row, column=col_nome).value
        proc = ws.cell(row=row, column=col_processo).value
        nome_s = str(nome).strip() if nome is not None else ""
        proc_s = str(proc).strip() if proc is not None else ""
        yield telefone, nome_s, proc_s


def _escolher_planilha_lemitti(wb, nome_aba: Optional[str]) -> Worksheet:
    """Primeira aba com cabeçalho típico em Z (POSSUI…WHATS…), ou aba explícita, ou primeira do livro."""
    if nome_aba and nome_aba.strip():
        if nome_aba.strip() not in wb.sheetnames:
            raise ValueError(
                f"Aba de enriquecimento '{nome_aba.strip()}' não encontrada. Abas: {list(wb.sheetnames)}"
            )
        return wb[nome_aba.strip()]

    for name in wb.sheetnames:
        ws = wb[name]
        z_key = _norm_header(str(ws.cell(row=1, column=COL_WA).value or ""))
        if "possui" in z_key and "what" in z_key:
            return ws

    return wb[wb.sheetnames[0]]


def mapa_telefone_whatsapp_por_processo_lemitti(
    ws: Worksheet,
    linha_cabecalho: int,
    linha_dados_inicio: int,
    col_processo: int,
) -> Dict[str, str]:
    """
    Percorre o ficheiro Lemitti: DD (U) + Telefone (V), só Z=1 WhatsApp.
    Retorna: chave_processo_normalizada -> telefone (primeiro por processo).
    """
    out: Dict[str, str] = {}
    for row in range(linha_dados_inicio, ws.max_row + 1):
        ddd = _only_digits(ws.cell(row=row, column=COL_DDD).value)
        tel = _only_digits(ws.cell(row=row, column=COL_TEL).value)
        if not _whatsapp_flag(ws.cell(row=row, column=COL_WA).value):
            continue
        if not ddd and not tel:
            continue
        telefone = f"{ddd}{tel}"
        if not telefone:
            continue
        proc = ws.cell(row=row, column=col_processo).value
        proc_s = str(proc).strip() if proc is not None else ""
        chave = _norm_processo_key(proc_s)
        if not chave:
            continue
        if chave not in out:
            out[chave] = telefone
    return out


def _escolher_planilha_fonte(wb, nome_aba: Optional[str]) -> Tuple[Worksheet, Layout]:
    if nome_aba and nome_aba.strip():
        if nome_aba.strip() not in wb.sheetnames:
            raise ValueError(
                f"Aba '{nome_aba.strip()}' não encontrada. Abas: {list(wb.sheetnames)}"
            )
        ws = wb[nome_aba.strip()]
        layout = _detect_layout(ws, 1)
        if layout is None:
            raise ValueError(
                f"Aba '{nome_aba.strip()}' não reconhecida (nem Lemitti U/V/Z nem TELEFONE_1…)."
            )
        return ws, layout

    for name in wb.sheetnames:
        ws = wb[name]
        layout = _detect_layout(ws, 1)
        if layout is not None:
            return ws, layout

    raise ValueError(
        "Nenhuma aba com layout reconhecido. "
        "Esperado: colunas TELEFONE_1… (PRC FINAL) ou Lemitti com POSSUI-WHATSAPP na coluna Z."
    )


@dataclass
class WhatsAppAbaConfig:
    caminho_arquivo: str
    nome_aba_fonte: Optional[str] = None
    nome_aba_destino: str = NOME_ABA_DESTINO_PADRAO
    linha_cabecalho: int = 1
    linha_dados_inicio: int = 2
    col_nome: Optional[str] = None
    col_processo: Optional[str] = None
    exige_flag_whatsapp: bool = False
    # Ficheiro separado só de enriquecimento Lemitti (U+V+Z); cruza com o FINAL pelo processo
    caminho_enriquecimento_lemitti: Optional[str] = None
    nome_aba_enriquecimento: Optional[str] = None
    linha_cabecalho_lemitti: int = 1
    linha_dados_inicio_lemitti: int = 2
    col_processo_lemitti: Optional[str] = None


def _col_letter_to_idx(letter: Optional[str]) -> Optional[int]:
    if not letter or not str(letter).strip():
        return None
    return column_index_from_string(str(letter).strip().upper())


def _aplicar_merge_final_com_lemitti(cfg: WhatsAppAbaConfig) -> Dict[str, object]:
    """
    Telefone vem do Excel Lemitti (U+V, Z=1). Nome e processo vêm do FINAL (aba PRC).
    """
    caminho_lem = cfg.caminho_enriquecimento_lemitti
    if not caminho_lem or not str(caminho_lem).strip():
        raise ValueError("caminho_enriquecimento_lemitti é obrigatório para o modo merge.")

    wb_lem = load_workbook(caminho_lem, data_only=True)
    try:
        ws_lem = _escolher_planilha_lemitti(wb_lem, cfg.nome_aba_enriquecimento)
        titulo_aba_lemitti = ws_lem.title
        hr_l = cfg.linha_cabecalho_lemitti
        ds_l = cfg.linha_dados_inicio_lemitti
        col_pl = _col_letter_to_idx(cfg.col_processo_lemitti)
        if col_pl is None:
            _, det_proc = _find_header_columns(ws_lem, hr_l)
            col_pl = det_proc
        if col_pl is None:
            raise ValueError(
                "No ficheiro Lemitti não foi encontrada coluna de número de processo. "
                "Use --col-processo-lemitti (letra Excel)."
            )
        mapa_tel = mapa_telefone_whatsapp_por_processo_lemitti(ws_lem, hr_l, ds_l, col_pl)
    finally:
        wb_lem.close()

    wb = load_workbook(cfg.caminho_arquivo, data_only=True)
    try:
        ws_src, layout = _escolher_planilha_fonte(wb, cfg.nome_aba_fonte)
        if layout != "prc_final":
            raise ValueError(
                "Com ficheiro Lemitti separado, o FINAL deve ter layout PRC (cabeçalho TELEFONE_1…). "
                f"Aba '{ws_src.title}' foi detectada como {layout!r}."
            )

        hr = cfg.linha_cabecalho
        ds = cfg.linha_dados_inicio
        col_nome = _col_letter_to_idx(cfg.col_nome)
        col_proc = _col_letter_to_idx(cfg.col_processo)
        if col_nome is None or col_proc is None:
            det_nome, det_proc = _find_header_columns(ws_src, hr)
            col_nome = col_nome or det_nome
            col_proc = col_proc or det_proc
        if col_nome is None or col_proc is None:
            raise ValueError(
                "No FINAL: não foi possível localizar Nome e Número de processo. "
                "Use --col-nome e --col-processo."
            )

        vistos: set = set()
        saida: List[Tuple[str, str, str]] = []

        for row in range(ds, ws_src.max_row + 1):
            proc = ws_src.cell(row=row, column=col_proc).value
            proc_s = str(proc).strip() if proc is not None else ""
            chave = _norm_processo_key(proc_s)
            if not chave or chave in vistos:
                continue
            telefone = mapa_tel.get(chave)
            if not telefone:
                continue
            vistos.add(chave)
            nome = ws_src.cell(row=row, column=col_nome).value
            nome_s = str(nome).strip() if nome is not None else ""
            saida.append((telefone, nome_s, proc_s))

        if cfg.nome_aba_destino in wb.sheetnames:
            wb.remove(wb[cfg.nome_aba_destino])
        ws_out = wb.create_sheet(cfg.nome_aba_destino)
        ws_out.append(["Telefone", "Nome", "Número de processo"])
        for telefone, nome, proc in saida:
            ws_out.append([telefone, nome, proc])

        wb.save(cfg.caminho_arquivo)

        return {
            "linhas_escritas": len(saida),
            "processos_unicos": len(vistos),
            "layout": "merge_lemitti_final",
            "aba_fonte": ws_src.title,
            "aba_lemitti": titulo_aba_lemitti,
            "processos_com_whatsapp_lemitti": len(mapa_tel),
            "aviso": None,
        }
    finally:
        wb.close()


def aplicar_aba_whatsapp_final(cfg: WhatsAppAbaConfig) -> Dict[str, object]:
    if cfg.caminho_enriquecimento_lemitti:
        return _aplicar_merge_final_com_lemitti(cfg)

    wb = load_workbook(cfg.caminho_arquivo, data_only=True)
    try:
        ws_src, layout = _escolher_planilha_fonte(wb, cfg.nome_aba_fonte)
        hr = cfg.linha_cabecalho
        ds = cfg.linha_dados_inicio

        col_nome = _col_letter_to_idx(cfg.col_nome)
        col_proc = _col_letter_to_idx(cfg.col_processo)

        if col_nome is None or col_proc is None:
            det_nome, det_proc = _find_header_columns(ws_src, hr)
            col_nome = col_nome or det_nome
            col_proc = col_proc or det_proc

        if col_nome is None or col_proc is None:
            raise ValueError(
                "Não foi possível localizar colunas de Nome e Número de processo. "
                "Use --col-nome e --col-processo se necessário."
            )

        vistos: set = set()
        saida: List[Tuple[str, str, str]] = []

        if layout == "lemitti_uvz":
            iterator = iter_linhas_lemitti_uvz(ws_src, ds, col_nome, col_proc)
        else:
            iterator = iter_linhas_prc_final(
                ws_src, hr, ds, col_nome, col_proc, cfg.exige_flag_whatsapp
            )

        for telefone, nome, proc in iterator:
            chave = _norm_processo_key(proc)
            if not chave:
                continue
            if chave in vistos:
                continue
            vistos.add(chave)
            saida.append((telefone, nome, proc))

        if cfg.nome_aba_destino in wb.sheetnames:
            wb.remove(wb[cfg.nome_aba_destino])
        ws_out = wb.create_sheet(cfg.nome_aba_destino)

        ws_out.append(["Telefone", "Nome", "Número de processo"])
        for telefone, nome, proc in saida:
            ws_out.append([telefone, nome, proc])

        wb.save(cfg.caminho_arquivo)

        aviso: Optional[str] = None
        if layout == "prc_final":
            _, poss = _parse_telefone_possui_cols(ws_src, hr)
            if not poss and not cfg.exige_flag_whatsapp:
                aviso = (
                    "Esta planilha não tem colunas POSSUI_WHATSAPP_n; "
                    "foi usado o primeiro TELEFONE_k preenchido por linha."
                )

        return {
            "linhas_escritas": len(saida),
            "processos_unicos": len(vistos),
            "layout": layout,
            "aba_fonte": ws_src.title,
            "aviso": aviso,
        }
    finally:
        wb.close()


def _norm_processo_key(proc: str) -> str:
    if not proc or not str(proc).strip():
        return ""
    return re.sub(r"\s+", "", str(proc).upper().strip())
